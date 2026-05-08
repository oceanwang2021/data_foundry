import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def parse_sql_schema(sql_file):
    with open(sql_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    tables = {}
    
    # Find all CREATE TABLE statements
    pattern = r'CREATE TABLE (\w+)\s*\((.*?)\)\s*ENGINE=InnoDB.*?;'
    matches = re.findall(pattern, content, re.DOTALL)
    
    for table_name, columns_text in matches:
        columns = []
        
        # Split by comma but be careful with nested commas in JSON etc
        lines = columns_text.strip().split('\n')
        
        for line in lines:
            line = line.strip()
            
            # Skip constraints (PRIMARY KEY, INDEX, etc.)
            if line.startswith('PRIMARY KEY') or line.startswith('INDEX') or \
               line.startswith('KEY') or line.startswith('CONSTRAINT') or \
               line.startswith(')'):
                continue
            
            # Parse column definition
            # Match: column_name type [options]
            col_match = re.match(r'^\s*(\w+)\s+(\w+(?:\(\d+(?:,\d+)?\))?)\s*(.*)', line)
            
            if col_match:
                col_name = col_match.group(1)
                col_type = col_match.group(2)
                col_options = col_match.group(3)
                
                # Add column info
                columns.append({
                    'name': col_name,
                    'type': col_type,
                    'options': col_options
                })
        
        tables[table_name] = columns
    
    return tables

def get_column_comment(col_name, table_name):
    """Generate comments based on column naming conventions"""
    comments = {
        #通用字段
        'id': '主键ID',
        'sort_order': '排序顺序',
        'created_at': '创建时间',
        'updated_at': '更新时间',
        'created_by': '创建人',
        'status': '状态',
        'name': '名称',
        'description': '描述',
        'title': '标题',
        '_business_date': '业务日期',
        'business_date': '业务日期',
        'request_id': '请求ID',
        'batch_id': '批次ID',
        'source_type': '源类型',
        'triggered_by': '触发者',
        'total_tasks': '总任务数',
        'completed_tasks': '已完成任务数',
        'failed_tasks': '失败任务数',
        'indicator_key': '指标键',
        'indicator_keys_json': '指标键JSON',
        'dimension_values_json': '维度值JSON',
        'indicator_values_json': '指标值JSON',
        'system_values_json': '系统值JSON',
        'row_binding_key': '行绑定键',
        'request_id': '请求ID',
        'parent_task_id': '父任务ID',
        'execution_mode': '执行模式',
        'schema_version': '模式版本',
        'plan_version': '计划版本',
        'date': '日期',
        'data_source': '数据源',
        'collection_policy': '采集策略',
        'processing_rule_drafts': '处理规则草稿',
        'collection_batches': '采集批次',
        'semantic_time_axis': '语义时间轴',
        'schema_json': '模式JSON',
        'scope_json': '范围JSON',
        'indicator_groups_json': '指标组JSON',
        'schedule_rules_json': '调度规则JSON',
        'collection_coverage_mode': '采集覆盖模式',
        'table_name': '表名',
        'wide_table_id': '宽表ID',
        'requirement_id': '需求ID',
        'parent_requirement_id': '父需求ID',
        'phase': '阶段',
        'schema_locked': '模式锁定',
        'owner': '所有者',
        'assignee': '负责人',
        'business_goal': '业务目标',
        'background_knowledge': '背景知识',
        'business_boundary': '业务边界',
        'delivery_scope': '交付范围',
        'data_update_enabled': '数据更新启用',
        'data_update_mode': '数据更新模式',
        'collection_policy': '集合策略',
        'snapshot_at': '快照时间',
        'snapshot_label': '快照标签',
        'coverage_mode': '覆盖模式',
        'is_current': '是否当前',
        'plan_version': '计划版本',
        'start_business_date': '开始业务日期',
        'end_business_date': '结束业务日期',
        'row_id': '行ID',
        'row_status': '行状态',
        'index_id': '索引ID',
        'indicator_group_id': '指标组ID',
        'indicator_group_name': '指标组名称',
        'name': '名称',
        'query': '查询语句',
        'narrow_row_json': '窄行JSON',
        'trigger_type': '触发类型',
        'started_at': '开始时间',
        'ended_at': '结束时间',
        'operator': '操作员',
        'output_ref': '输出引用',
        'log_ref': '日志引用',
        'can_rerun': '是否可重跑',
        'invalidated_reason': '失效原因',
        'confidence': '置信度',
        'batch_id': '批次ID',
        'task_group_id': '任务组ID',
        'backfill_request_id': '回填请求ID',
        'schedule_rule_id': '调度规则ID',
        'backfill_request_id': '回填请求ID',
        'group_kind': '组类型',
        'partition_type': '分区类型',
        'partition_key': '分区键',
        'partition_label': '分区标签',
        'document_count': '文档数量',
        'last_updated': '最后更新',
        'category': '分类',
        'expression': '表达式',
        'sample_issue': '样本问题',
        'indicator_bindings_json': '指标绑定JSON',
        'filling_config_json': '填充配置JSON',
        'mode': '模式',
        'scenario_rigour': '场景严格性',
        'condition_expr': '条件表达式',
        'action_text': '动作文本',
        'enabled': '是否启用',
        'enabled': '是否启用',
        'document_count': '文档数量',
        'last_updated': '最后更新',
        'category': '分类',
        'expression': '表达式',
        'sample_issue': '样本问题',
        'indicator_bindings_json': '指标绑定JSON',
        'filling_config_json': '填充配置JSON',
        'mode': '模式',
        'scenario_rigour': '场景严格性',
        'condition_expr': '条件表达式',
        'action_text': '动作文本',
        'enabled': '是否启用',
        'enabled': '是否启用',
        'business_background': '业务背景',
        'business_goal': '业务目标',
        'background_knowledge': '背景知识',
        'business_boundary': '业务边界',
        'delivery_scope': '交付范围',
        'data_update_enabled': '数据更新启用',
        'data_update_mode': '数据更新模式',
        'processing_rule_drafts': '处理规则草稿',
        'collection_policy': '采集策略',
        'schema_locked': '模式锁定',
        'semantic_time_axis': '语义时间轴',
        'collection_coverage_mode': '采集覆盖模式',
        'record_count': '记录数',
        'batch_id': '批次ID',
        'snapshot_label': '快照标签',
        'coverage_mode': '覆盖模式',
        'is_current': '是否当前',
        'plan_version': '计划版本',
        'triggered_by': '触发者',
        'start_business_date': '开始业务日期',
        'end_business_date': '结束业务日期',
        'row_id': '行ID',
        'row_status': '行状态',
        'origin': '来源',
        'reason': '原因',
        'partition_type': '分区类型',
        'partition_key': '分区键',
        'partition_label': '分区标签',
        'total_tasks': '总任务数',
        'completed_tasks': '已完成任务数',
        'failed_tasks': '失败任务数',
        'group_kind': '组类型',
    }
    
    # Remove trailing underscore if present
    clean_name = col_name.rstrip('_')
    
    if clean_name in comments:
        return comments[clean_name]
    
    # Default comment based on naming
    if col_name.endswith('_id'):
        return f'{col_name[:-3]}ID'
    elif col_name.endswith('_at'):
        return f'{col_name[:-3]}时间'
    elif col_name.endswith('_json'):
        return f'{col_name[:-5]}JSON'
    elif col_name.endswith('_date'):
        return f'{col_name[:-5]}日期'
    elif col_name.endswith('_key'):
        return f'{col_name[:-4]}键'
    elif col_name.endswith('_name'):
        return f'{col_name[:-5]}名称'
    elif col_name.endswith('_type'):
        return f'{col_name[:-5]}类型'
    elif col_name.endswith('_count'):
        return f'{col_name[:-6]}数量'
    elif col_name.endswith('_enabled'):
        return f'{col_name[:-8]}启用'
    elif col_name.endswith('_status'):
        return f'{col_name[:-7]}状态'
    elif col_name.endswith('_version'):
        return f'{col_name[:-8]}版本'
    
    return col_name

def map_sql_type_to_excel(sql_type):
    """Map SQL types to Excel display format"""
    type_mapping = {
        'varchar': 'VARCHAR',
        'int': 'INT',
        'bigint': 'BIGINT',
        'tinyint': 'TINYINT',
        'smallint': 'SMALLINT',
        'mediumint': 'MEDIUMINT',
        'decimal': 'DECIMAL',
        'double': 'DOUBLE',
        'float': 'FLOAT',
        'real': 'REAL',
        'numeric': 'NUMERIC',
        'integer': 'INTEGER',
        'bool': 'BOOLEAN',
        'boolean': 'BOOLEAN',
        'date': 'DATE',
        'datetime': 'DATETIME',
        'timestamp': 'TIMESTAMP',
        'time': 'TIME',
        'year': 'YEAR',
        'text': 'TEXT',
        'tinytext': 'TINYTEXT',
        'mediumtext': 'MEDIUMTEXT',
        'longtext': 'LONGTEXT',
        'blob': 'BLOB',
        'tinyblob': 'TINYBLOB',
        'mediumblob': 'MEDIUMBLOB',
        'longblob': 'LONGBLOB',
        'char': 'CHAR',
        'binary': 'BINARY',
        'varbinary': 'VARBINARY',
        'json': 'JSON',
    }
    
    sql_type_lower = sql_type.lower()
    for sql_pattern, excel_type in type_mapping.items():
        if sql_pattern in sql_type_lower:
            return excel_type
    
    return sql_type.upper()

def generate_excel(tables, output_file):
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    for table_name, columns in tables.items():
        # Create sheet with table name (max 31 chars for Excel)
        sheet_name = table_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        
        # Header row
        headers = ['字段名', '字段注释', '字段类型']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Column data
        for row_idx, col in enumerate(columns, 2):
            ws.cell(row=row_idx, column=1, value=col['name'])
            ws.cell(row=row_idx, column=2, value=get_column_comment(col['name'], table_name))
            ws.cell(row=row_idx, column=3, value=map_sql_type_to_excel(col['type']))
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
    
    # Save workbook
    wb.save(output_file)
    print(f"Excel file generated: {output_file}")

if __name__ == '__main__':
    sql_file = r'd:\trae\data_foundry\data_foundry\db\mysql\backend\002_full_schema.sql'
    output_file = r'd:\trae\data_foundry\data_foundry\db\mysql\backend\data_foundry_schema.xlsx'
    
    tables = parse_sql_schema(sql_file)
    generate_excel(tables, output_file)
    
    print(f"\nTotal tables processed: {len(tables)}")
    for table_name in tables.keys():
        print(f"  - {table_name}: {len(tables[table_name])} columns")
